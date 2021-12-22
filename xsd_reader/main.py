import xmlschema
from openpyxl import load_workbook
import ntpath
from openpyxl.styles import PatternFill

class XsdReader:
    def __init__(self,ex_path,sheet, letter, end_num,schema_path,export):
        self.ex_path = ex_path
        self.sheet = sheet
        self.letter = letter
        self.end_num = end_num
        self.schema = xmlschema.XMLSchema(schema_path)
        self.export = export

    def index_builder(self, path_ex):
        path = path_ex

        listo = path.split('/')
        fir = listo[0]
        empt = True
        path_found = False
        for elem in listo[1:]:
            path = fir
            path = path + '/'
            if empt:
                for i in self.schema.namespaces.keys():
                    if not i == '':
                        path = path + i + ':' + elem
                    else:
                        path = path + elem
                    # print(path)
                    res = self.schema.findall(path, namespaces=self.schema.namespaces)
                    if not res == []:
                        fir = path
                        # print('path found')
                        path_found = True
                        break
                    else:
                        path = fir
                        path = path + '/'
                        path_found = False
        print(fir)
        return path_found

    def check(self):
        wb = load_workbook(self.ex_path)

        sheet = wb[self.sheet]
        schema = self.schema


        for i in sheet[f'{self.letter}2':f'{self.letter}{self.end_num}']:
            print(f'Num {i}')
            for cell in i:
                if self.export:
                    cell.value = 'export/'+cell.value
                if '.' in cell.value:
                    cell.value = cell.value.replace('.', '/')
                if ' ' in cell.value:
                    cell.value = cell.value.replace(' ', '')
                if '\n' in cell.value:
                    path_one, path_two = cell.value.split('\n')
                    print(path_one,path_two)
                    paths = [path_one,path_two]
                    for path in paths:
                        path_res = self.index_builder(path)
                        if path_res:
                            print("path_found")
                            cell.fill = PatternFill("solid", start_color="00FFFF99")
                        else:
                            print("path_not_found")
                            cell.fill = PatternFill("solid", start_color="0099CC00")
                else:
                    path = cell.value
                    path_res = self.index_builder(path)
                    print(path)
                    if path:
                        print("path_found")
                        cell.fill = PatternFill("solid", start_color="0099CC00")
                    else:
                        print("path_not_found")
                        cell.fill = PatternFill("solid", start_color="00FFFF99")
        wb.save(r"C:\Users\mixaz\Desktop\Необходимые файлы\testo.xlsx")
        

check = XsdReader(r'C:\Users\mixaz\Downloads\Telegram_Desktop\test.xlsx','notification','E',850,
                  r'C:\Users\mixaz\Desktop\Необходимые файлы\Схемы 11.3 итерация 1\fcsExport.xsd',export=True)

check.check()