import xmlschema
from openpyxl import load_workbook
import ntpath
from openpyxl.styles import PatternFill


schema = xmlschema.XMLSchema(r'C:\Users\mixaz\Desktop\Необходимые файлы\Схемы 11.3 итерация 1\fcsExport.xsd')
# namespaces = {'oos': 'http://zakupki.gov.ru/oos/types/1','':'http://zakupki.gov.ru/oos/export/1'}
#
print(schema.findall('export/fcsNotificationOKOU/oos:id',schema.namespaces))
print(schema.namespaces.keys())


# wb = load_workbook(r'C:\Users\mixaz\Desktop\описание маппинга 223 ФЗ.xlsx')
#
# sheet = wb['PROTOCOLS']
# if '\n' in sheet['E7'].value:
#     path1,path2 =  sheet['E7'].value.split('\n')
#     print(path1, '\n', path2)
#     print(ntpath.dirname(path1))
#     print(ntpath.basename(path1))
# else:
#     print('123')


path = 'export/epNotificationEZK/notificationInfo/customerRequirementsInfo/customerRequirementInfo/contractConditionsInfo/maxPriceInfo/maxPrice'

listo = path.split('/')
print(listo)
fir = listo[0]
empt = True
for elem in listo[1:]:
    path = fir
    path = path + '/'
    if empt:
        for i in schema.namespaces.keys():
            if not i == '':
                path = path + i + ':' + elem
            else:
                path = path + elem
            print(path)
            res = schema.findall(path, namespaces= schema.namespaces)
            if not res == []:
                fir = path
                print('path found')
                break
            else:
                path = fir
                path = path + '/'
